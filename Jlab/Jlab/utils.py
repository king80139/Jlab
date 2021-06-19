def Read_Arg_(username, prname, arguement, isind=0):  # Backbone Dictionary 스프레드 시트의 ProjectLibrary(recent)시트를 기반으로,
    # argument에 들어갈 명령어의 참조파일, input파일, output파일을 리턴하는 함수입니다.

    #창민이 버전
    #Lib = pd.read_excel(os.path.join(os.getcwd(), "media", username, prname, "jlabLibrary.xlsx")).fillna(
    #    "")  # worksheet를 pandas DataFrame으로 변환해줍니다.
    Lib = Read_Sheet_(username, prname,'ProjectLibrary(recent)')  # worksheet를 pandas DataFrame으로 변환해줍니다.

    target_row = Lib.loc[
        Lib["*함수명/ parameter이름"] == "*" + arguement, ["Reference File (Information)", "Input File/Information",
                                                      "Output file", "Action Status"]]
    # 원하는 함수가 쓰여있는 행을 찾고, 이 행에서 참조파일, input 파일, output 파일을 가져옵니다.
    if isind==0:
        target_row = target_row[target_row["Action Status"]!=""]
        target_row = target_row[target_row["Action Status"] > 0]
        target_row = target_row[target_row["Action Status"] == target_row["Action Status"].min()]
    else:pass
    if len(target_row) == 0:
        print("실행가능한 함수가 없습니다. Backbone Dictionary의 Action Status를 확인하십시오. 에러메세지가 발생합니다.")
        return None

    else:
        refFile_info, Input_info, Output = target_row[target_row.columns[0]].values[0], \
                                           target_row[target_row.columns[1]].values[0], \
                                           target_row[target_row.columns[2]].values[0]

        return [refFile_info, Input_info, Output]  # 참조파일, input 파일, output 파일을 차례로 리턴해줍니다.

def option_finder(arg):
    sht = Read_Sheet_("ProjectLibrary(recent)")
    underrow=sht.shift(-1)[
        sht["*함수명/ parameter이름"]==f"*{arg}"].applymap(
            lambda x: None if x=="" else x
            ).dropna(axis=1).values[0]
    if underrow[0] == "option":
        option = dict()
        for i,o in enumerate(underrow[1:]):
            option[f"option_{i+1}"] = o
        return option
    else :
        print("no option")
        pass


class OS:
    def encoding(self):
        import platform
        if platform.system() == "Darwin" or platform.system() == "Linux":
            encoding = "cp949"
        elif platform.system() == "Windows":
            encoding = "utf-8-sig"
        else:
            encoding = "cp949"
        return encoding


def Read_Sheet_(username, prname, sheet_name):  # Backbone Dictionary 스프레드 시트를 기반으로,
    # sheet_name에 들어갈 시트를 리턴하는 함수입니다.
    import pandas as pd
    import os
    try:
        #worksheet = pd.read_excel(os.path.join(os.getcwd(), "media", username, prname, "JlabMiner library Backbone Dictionary.xlsx"), sheet_name=sheet_name).fillna("")# 창민이 버전
        if (username is None) or (prname is None):
            worksheet = pd.read_excel("JlabMiner library Backbone Dictionary.xlsx",  #내 버전 # Backbone Dictionary가 있는 경로를 써줘야 합니다.
                                  sheet_name=sheet_name).fillna("")  # 창민아 여기에 서버에 올린 Library 주소를 쓰면 될거야.
        elif (username is not None) and (prname is not None):
            worksheet = pd.read_excel(username+"/"+prname+"/"+"JlabMiner library Backbone Dictionary.xlsx",
                                      # 내 버전 # Backbone Dictionary가 있는 경로를 써줘야 합니다.
                                      sheet_name=sheet_name).fillna("")  # 창민아 여기에 서버에 올린 Library 주소를 쓰면 될거야.
        else :
            print("username, prname이 제대로 적히지 않았습니다.")
    except FileNotFoundError as e:
        if os.path.splitext(sheet_name) == ".csv":
            opsys = OS()
            encoding = opsys.encoding()
            worksheet = pd.read_csv(sheet_name, encoding=encoding).fillna("")
        elif os.path.splitext(sheet_name) == ".xlsx":
            worksheet = pd.read_excel(sheet_name).fillna("")
        elif os.path.splitext(sheet_name) == ".pkl":
            worksheet = pd.read_pickle(sheet_name).fillna("")
        else:
            raise e

    return worksheet


def sequential_run():
    from openpyxl import load_workbook
    from .Jlab_Text_Cleaning_Functions import Delete_Messages, Delete_Characters, Delete_Characters_by_Dic, Delete_Overlapped_Messages, Delete_StandardStopwords, Replace_Texts_in_Messages, Replace_Texts_by_Dic, Frequency_Analysis, Make_Cooccurrence_Table
    from .Calculate_BD import Calculate_BD
    from .Draw_Static_3D_plot_Space import draw_static_3d_plot_space
    from .plots import draw_snplot, draw_WordCloud

    wb = load_workbook(filename="JlabMiner library Backbone Dictionary.xlsx")
    ws = wb.get_sheet_by_name("ProjectLibrary(recent)")
    sht = Read_Sheet_("ProjectLibrary(recent)")
    sht_ = sht.copy()
    sht_ = sht_[
               ["*함수명/ parameter이름", "Action Status"]
           ][5:].applymap(lambda cell: None if cell == "" else cell).dropna()
    sht_ = sht_[sht_["Action Status"] > 0].sort_values(by="Action Status", ascending=True)
    seq = ["".join([i.replace("*", ""), "()"]) for i in sht_.to_dict(orient="list")["*함수명/ parameter이름"]]
    action_stat = [actst for actst in sht_.to_dict(orient="list")["Action Status"]]
    last_func = len(seq)
    for i, func in enumerate(seq):
        try:
            print(f"execute {func}...")
            res = eval(func)
            xl_idx = sht[(sht["*함수명/ parameter이름"] == "".join(["*", func.replace("()", "")])) & (
                        sht["Action Status"] == action_stat[i])].index[0] + 2
            xl_col = list(sht.columns).index("Action Status") + 1
            ws.cell(row=xl_idx, column=xl_col).value = -1
            wb.save("JlabMiner library Backbone Dictionary.xlsx")
            if i + 1 == last_func:
                return res
            else:
                pass
        except Exception as e:
            if i + 1 == 1:
                print(f"{seq[i]} raised a some kind of error.")

            else:
                print(f"processed up to {seq[i - 1]}, but {seq[i]} raised a some kind of error.")
            print(e)
            return res



def import_dataframe(input_name):
    import pandas as pd

    opsys = OS()
    encoding = opsys.encoding()

    if ".csv" in input_name:
        try:
            dataframe = pd.read_csv(input_name, encoding='utf-8')
        except:
            dataframe = pd.read_csv(input_name, encoding='cp949')
        return dataframe
    elif ".xlsx" in input_name:
        dataframe = pd.read_excel(input_name)
        return dataframe
    elif ".pkl" in input_name:
        dataframe = pd.read_pickle(input_name)
        return dataframe
    else:
        if type(input_name) == dict:
            return pd.DataFrame(input_name)
        elif type(input_name) == pd.core.frame.DataFrame:
            return input_name
        else:
            print("Not supported data type")



def export_dataframe(dataframe, output_name):
    import os
    name, ext = os.path.splitext(output_name)

    opsys = OS()
    encoding = opsys.encoding()

    def export_(df, out_name, enc):

        try:
            df.to_csv(out_name, encoding=enc, index=False)
        except UnicodeEncodeError:
            df.to_excel(out_name, index=False)
            print("csv파일로 내보내는데 오류가 있어 .xlsx파일로 내보냅니다.")

    if ".csv" in output_name:
        export_(dataframe, output_name, encoding)

    elif ".xlsx" in output_name:
        dataframe.to_excel(output_name, index=False)

    elif (name != "") & (ext == ""):
        output_name = "".join([name, ".csv"])
        export_(dataframe, output_name, encoding)

    elif output_name is None:
        output_name = input("결과물 이름이 정해지지 지지 않았습니다. 결과물의 이름을 기입해주십시오(확장자 제외) : ")
        output_name = "".join([output_name, ".csv"])
        export_(dataframe, output_name, encoding)
    else:
        # export_(dataframe, output_name, encoding)
        dataframe.to_pickle("".join([name, ".pkl"]))
        print("오류 발생, 일단 pickle파일로 export.")
        pass


__all__ = ['OS', 'Read_Arg_', 'Read_Sheet_', 'option_finder' ,'import_dataframe', 'export_dataframe', 'sequential_run']
